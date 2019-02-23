
# coding: utf-8

# In[1]:


import pandas as pd


# In[2]:


XTrain = pd.read_csv("X_train.csv")
ytrain = pd.read_csv("y_train.csv")
XTest = pd.read_csv("X_test.csv")
ytest = pd.read_csv("y_test.csv")


# In[3]:


from sklearn.ensemble import RandomForestClassifier
clf = RandomForestClassifier(n_estimators = 1000, max_depth = 3, random_state = 0)
clf.fit(XTrain, ytrain)


# In[4]:


clf.predict(XTest)


# In[5]:


scoreTrain = clf.score(XTrain, ytrain)
scoreTest = clf.score(XTest, ytest)


# In[6]:


result_single = pd.DataFrame({'Tree Depth': 3, 'Training Score': [scoreTrain], 'Test Score': [scoreTest]} )
print(result_single)


# For multiple tree depth:

# In[7]:


import warnings
warnings.filterwarnings("ignore")


# In[8]:


results = pd.DataFrame(columns = ['TreeDepth', 'Training Score', 'Test Score'])
for treeDepth in range(1,11):
    clf = RandomForestClassifier(n_estimators = 1000, max_depth = treeDepth, random_state = 0)
    clf.fit(XTrain, ytrain)
    clf.predict(XTest)
    scoreTrain = clf.score(XTrain, ytrain)
    scoreTest = clf.score(XTest, ytest)
    results.loc[treeDepth] = [treeDepth, scoreTrain, scoreTest]

print(results.head(11))


# In[9]:


from matplotlib import pyplot as plt
ax = plt.plot(results.TreeDepth, results['Training Score'], results['Test Score'])
ay = plt.title("Score chart for Random Forest Classifier")
a = plt.xlabel("Tree Depth")
b = plt.ylabel("Scores")
c = plt.legend()


# In[10]:


import numpy as np

importances = clf.feature_importances_
std = np.std([tree.feature_importances_ for tree in clf.estimators_], axis = 0)
indices = np.argsort(importances)[::-1]

print("Feature Rankings:")

for f in range(XTrain.shape[1]):
    print("%d. Feature %d (%f)" % (f+1, indices[f], importances[indices[f]]))

plt.figure()
plt.title("Feature Importances:")
plt.bar(range(XTrain.shape[1]), importances[indices])
plt.xticks(range(XTrain.shape[1]), indices)
plt.xlim([-1, XTrain.shape[1]])
plt.xlabel("Feature Number (in decending order of importance)")
plt.ylabel("Ranking Value")
plt.show()


# <b>Source:</b>
# <br>https://scikit-learn.org/stable/auto_examples/ensemble/plot_forest_importances.html

# In[11]:


path = "Results.xlsx"
from openpyxl import load_workbook

wb = load_workbook(path)
if "Task 4" not in wb.sheetnames:
    wb.create_sheet("Task 4_Random Forrest")
    sheet = wb['Task 4_Random Forrest']
else: sheet = wb['Task 4_Random Forrest']

from openpyxl.utils.dataframe import dataframe_to_rows
rows = dataframe_to_rows(results, index = False)

for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
         sheet.cell(row=r_idx, column=c_idx, value=value)

wb.save(path)

