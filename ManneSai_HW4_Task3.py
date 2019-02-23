
# coding: utf-8

# In[1]:


import pandas as pd


# In[2]:


df_XTrain = pd.read_csv("X_train.csv")
df_ytrain = pd.read_csv("y_train.csv")
df_XTest = pd.read_csv("X_test.csv")
df_ytest = pd.read_csv("y_test.csv")

file_input = "cancerNormalized.csv"
df = pd.read_csv(file_input)


# In[3]:


import warnings
warnings.filterwarnings("ignore")


# Minkowski with 3 Nearest Neighbors and p =2

# In[4]:


from sklearn.neighbors import KNeighborsClassifier
knn = KNeighborsClassifier(n_neighbors = 3, p=2, metric = 'minkowski')
knn.fit(df_XTrain, df_ytrain)


# In[5]:


knn.predict(df_XTest)


# In[6]:


scoreTrain = knn.score(df_XTrain, df_ytrain)
print(scoreTrain)


# In[7]:


scoreTest = knn.score(df_XTest, df_ytest)
print(scoreTest)


# Minkowski with 5 Nearest Neighbors with p=2

# In[8]:


from sklearn.neighbors import KNeighborsClassifier
knn = KNeighborsClassifier(n_neighbors = 5, p=2, metric = 'minkowski')
knn.fit(df_XTrain, df_ytrain)


# In[9]:


knn.predict(df_XTest)


# In[10]:


scoreTrain_M5= knn.score(df_XTrain, df_ytrain)
print(scoreTrain_M5)
scoreTest_M5 = knn.score(df_XTest, df_ytest)
print(scoreTest_M5)


# Euclidean with 5 nearest neighbors and p=2

# In[11]:


from sklearn.neighbors import KNeighborsClassifier
knn_Eucledian_5 = KNeighborsClassifier(n_neighbors=5, p=2, metric = 'euclidean')
knn_Eucledian_5.fit(df_XTrain, df_ytrain)


# In[12]:


knn_Eucledian_5.predict(df_XTest)


# In[13]:


scoreTrain_E5 = knn_Eucledian_5.score(df_XTrain, df_ytrain)
print(scoreTrain_E5)
scoreTest_E5 = knn_Eucledian_5.score(df_XTest, df_ytest)
print(scoreTest_E5)


# Euclidean with 3 nearest neighbors and p=2

# In[14]:


from sklearn.neighbors import KNeighborsClassifier
knn_Eucledian_3= KNeighborsClassifier(n_neighbors=3, p=2, metric = 'euclidean')
knn_Eucledian_3.fit(df_XTrain, df_ytrain)


# In[15]:


knn_Eucledian_3.predict(df_XTest)


# In[16]:


scoreTrain_E3 = knn_Eucledian_3.score(df_XTrain, df_ytrain)
print(scoreTrain_E3)
scoreTest_E3 = knn_Eucledian_3.score(df_XTest, df_ytest)
print(scoreTest_E3)


# In[17]:


df_table = pd.DataFrame(
    {
        'Labels':['Training Score','Test Score'],
        'Euclidean 3N' :[scoreTrain_E3, scoreTest_E3],
        'Euclidean 5N' :[scoreTrain_E5, scoreTest_E5],
        'Minkowski 3N' :[scoreTrain, scoreTest],
        'Minkowski 5N' :[scoreTrain_M5, scoreTest_M5]
    }, index = None
)


# In[18]:


print(df_table)


# In[19]:


path = "Results.xlsx"
from openpyxl import load_workbook

wb = load_workbook(path)
if "Task 3_KNN" not in wb.sheetnames:
    wb.create_sheet("Task 3_KNN")
sheet = wb['Task 3_KNN']

from openpyxl.utils.dataframe import dataframe_to_rows
rows = dataframe_to_rows(df_table, index = False)

for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
         sheet.cell(row=r_idx, column=c_idx, value=value)

wb.save(path)

