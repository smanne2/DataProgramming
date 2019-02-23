
# coding: utf-8

# In[1]:


import pandas as pd


# In[2]:


df_XTrain = pd.read_csv("X_train.csv")
df_ytrain = pd.read_csv("y_train.csv")
df_XTest = pd.read_csv("X_test.csv")
df_ytest = pd.read_csv("y_test.csv")

file_input = "cancerNormalized.csv"
df = pd.read_csv(file_input)


# In[3]:


df['RatioA1A3'] = df['area1']/df['area3']

df['RatioP1P3'] = df['perimeter1']/df['perimeter3']


# In[4]:


import numpy as np
np.unique(df.Class)


# In[5]:


df = df.fillna(0)


# In[6]:


list(enumerate(np.unique(df.Class)))


# In[7]:


class_mapping = {label: idx for idx, label in enumerate(np.unique(df.Class))}
print(class_mapping)


# In[8]:


df['Class'] = df['Class'].map(class_mapping)
print(df.head())


# In[9]:


from sklearn.preprocessing import MinMaxScaler


# In[10]:


scaler = MinMaxScaler(feature_range=(0,2))
new_df = scaler.fit_transform(df)
norm_df = pd.DataFrame(data = new_df[:,:], columns = df.columns)
from sklearn.model_selection import train_test_split
X,y = norm_df.iloc[:,1:], norm_df['Class']
df_XTrain, df_XTest, df_ytrain, df_ytest = train_test_split(X,y, test_size = (1/3), random_state = 0)


# In[11]:


import warnings
warnings.filterwarnings("ignore")


# Minkowski with 3 Nearest Neighbors and p =2

# In[12]:


from sklearn.neighbors import KNeighborsClassifier
knn = KNeighborsClassifier(n_neighbors = 3, p=2, metric = 'minkowski')
knn.fit(df_XTrain, df_ytrain)


# In[13]:


knn.predict(df_XTest)


# In[14]:


scoreTrain = knn.score(df_XTrain, df_ytrain)
print(scoreTrain)


# In[15]:


scoreTest = knn.score(df_XTest, df_ytest)
print(scoreTest)


# Minkowski with 5 Nearest Neighbors with p=2

# In[16]:


from sklearn.neighbors import KNeighborsClassifier
knn = KNeighborsClassifier(n_neighbors = 5, p=2, metric = 'minkowski')
knn.fit(df_XTrain, df_ytrain)


# In[17]:


knn.predict(df_XTest)


# In[18]:


scoreTrain_M5= knn.score(df_XTrain, df_ytrain)
print(scoreTrain_M5)
scoreTest_M5 = knn.score(df_XTest, df_ytest)
print(scoreTest_M5)


# Euclidean with 5 nearest neighbors and p=2

# In[19]:


from sklearn.neighbors import KNeighborsClassifier
knn_Eucledian_5 = KNeighborsClassifier(n_neighbors=5, p=2, metric = 'euclidean')
knn_Eucledian_5.fit(df_XTrain, df_ytrain)


# In[20]:


knn_Eucledian_5.predict(df_XTest)


# In[21]:


scoreTrain_E5 = knn_Eucledian_5.score(df_XTrain, df_ytrain)
print(scoreTrain_E5)
scoreTest_E5 = knn_Eucledian_5.score(df_XTest, df_ytest)
print(scoreTest_E5)


# Euclidean with 3 nearest neighbors and p=2

# In[22]:


from sklearn.neighbors import KNeighborsClassifier
knn_Eucledian_3= KNeighborsClassifier(n_neighbors=3, p=2, metric = 'euclidean')
knn_Eucledian_3.fit(df_XTrain, df_ytrain)


# In[23]:


knn_Eucledian_3.predict(df_XTest)


# In[24]:


scoreTrain_E3 = knn_Eucledian_3.score(df_XTrain, df_ytrain)
print(scoreTrain_E3)
scoreTest_E3 = knn_Eucledian_3.score(df_XTest, df_ytest)
print(scoreTest_E3)


# In[25]:


df_table = pd.DataFrame(
    {
        'Labels':['Training Score','Test Score'],
        'Euclidean 3N' :[scoreTrain_E3, scoreTest_E3],
        'Euclidean 5N' :[scoreTrain_E5, scoreTest_E5],
        'Minkowski 3N' :[scoreTrain, scoreTest],
        'Minkowski 5N' :[scoreTrain_M5, scoreTest_M5]
    }, index = None
)


# In[26]:


print(df_table)


# In[27]:


path = "Results.xlsx"
from openpyxl import load_workbook

wb = load_workbook(path)
if "Task 5_KNN" not in wb.sheetnames:
    wb.create_sheet("Task 5_KNN")
sheet = wb['Task 5_KNN']

from openpyxl.utils.dataframe import dataframe_to_rows
rows = dataframe_to_rows(df_table, index = False)

for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
         sheet.cell(row=r_idx, column=c_idx, value=value)

wb.save(path)

